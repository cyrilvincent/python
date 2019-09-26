import openpyxl as xl

wb = xl.load_workbook("data/house.xlsx")
sheet = wb.worksheets[0]
#sheet = wb["house"]
#print(sheet["A1"].value)
loyers=[]
surfaces=[]
for i in range(2, sheet.max_row + 1):
    loyers.append(sheet.cell(row=i, column=1).value)
    surfaces.append(sheet.cell(row=i, column=2).value)
wb.close()
print(list(zip(loyers, surfaces)))
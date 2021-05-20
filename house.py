import argparse
import os
import csv
import sys
import pickle
import xml.dom.minidom as dom
import json
import openpyxl

class HouseSuper:

    def __init__(self, path):
        self.path = path
        self.loyers = []
        self.surfaces = []
        self.loyers_per_m2 = []

    def min_max_avg(self, l):
        min = max = l[0]
        sum = 0
        for value in l:
            sum += value
            if value < min:
                min = value
            elif value > max:
                max = value
        return min, max, sum / len(l)

    def compute_loyers(self):
        return self.min_max_avg(self.loyers)

    def compute_surfaces(self):
        return self.min_max_avg(self.surfaces)

    def compute_loyers_per_m2(self):
        return self.min_max_avg(self.loyers_per_m2)


class HouseCsv(HouseSuper):

    def __init__(self, path, separator=","):
        super().__init__(path)
        self.separator = separator

    def load(self):
        with open(self.path) as f:
            reader = csv.DictReader(f, delimiter=self.separator)
            for row in reader:
                loyer = int(row["loyer"])
                surface = int(row["surface"])
                self.loyers.append(loyer)
                self.surfaces.append(surface)
                self.loyers_per_m2.append(loyer / surface)


class HousePickle(HouseSuper):

    def load(self):
        with open(self.path, "rb") as f:
            self.loyers, self.surfaces, self.loyers_per_m2 = pickle.load(f)

    def save(self):
        with open(self.path, "wb") as f:
            pickle.dump((self.loyers, self.surfaces, self.loyers_per_m2), f)


class HouseXml(HouseSuper):

    def load(self):
        doc = dom.parse(self.path)
        l = doc.getElementsByTagName("house")
        for node in l:
            loyer = int(node.getAttribute("loyer"))
            surface = int(node.getAttribute("surface"))
            self.loyers.append(loyer)
            self.surfaces.append(surface)
            self.loyers_per_m2.append(loyer / surface)

    def save(self):
        doc = dom.Document()
        houses_node = doc.createElement("houses")
        doc.appendChild(houses_node)
        for index, value in enumerate(self.loyers):
            house_node = doc.createElement("house")
            house_node.setAttribute("loyer", str(value))
            house_node.setAttribute("surface", str(self.surfaces[index]))
            houses_node.appendChild(house_node)
        with open(self.path, "w") as f:
            doc.writexml(f, addindent="\t", newl="\n" )

class HouseJson(HouseSuper):

    def load(self):
        with open(self.path) as f:
            dicos = json.load(f)
            for row in dicos:
                self.loyers.append(row["loyer"])
                self.surfaces.append(row["surface"])
                self.loyers_per_m2.append(row["loyer"] / row["surface"])

    def save(self):
        dicos = []
        for index, value in enumerate(self.loyers):
            row = {"loyer": value, "surface": self.surfaces[index]}
            dicos.append(row)
        with open(self.path, "w") as f:
            json.dump(dicos, f)


class HouseExcel(HouseSuper):

    def load(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]
        for i in range(2, sheet.max_row):
            loyer = sheet.cell(i,1).value
            surface = sheet.cell(i,2).value
            self.loyers.append(loyer)
            self.surfaces.append(surface)
            self.loyers_per_m2.append(loyer / surface)
        workbook.close()

    def save(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]
        for index, value in enumerate(self.loyers):
            sheet.cell(index + 2, 1).value = value
            sheet.cell(index + 2, 2).value = value
        workbook.save(self.path.replace(".xslx", "new.xlsx"))
        workbook.close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("path", help="Fichier à ouvrir")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbeux")
    parser.add_argument("-t", "--type", help="Type de fichier à ouvrir : csv, json, xml, xlsx")
    parser.add_argument("-s", "--separator", help="Séparateur des fichiers CSV")
    args = parser.parse_args()
    print(f"Loading {args.path}")
    if args.verbose:
        print(f"Current directory {os.getcwd()}")
        print(f"File type: {args.type}")
        if args.type == "csv":
            print(f"File separator: {args.separator}")
    if args.type == "csv":
        housecsv = HouseCsv(args.path, args.separator)
        housecsv.load()
        if args.verbose:
            min, max, avg = housecsv.compute_loyers()
            print(f"Loyers min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housecsv.compute_surfaces()
            print(f"Surfaces min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housecsv.compute_loyers_per_m2()
            print(f"Loyers par m² min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
        else:
            _, _, avg = housecsv.compute_loyers_per_m2()
            print(f"{avg:.2f}")
    elif args.type == "pickle":
        housepickle = HousePickle(args.path)
        housepickle.load()
        if args.verbose:
            min, max, avg = housepickle.compute_loyers()
            print(f"Loyers min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housepickle.compute_surfaces()
            print(f"Surfaces min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housepickle.compute_loyers_per_m2()
            print(f"Loyers par m² min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
        else:
            _, _, avg = housepickle.compute_loyers_per_m2()
            print(f"{avg:.2f}")
    elif args.type == "xml":
        housexml = HouseXml(args.path)
        housexml.load()
        if args.verbose:
            min, max, avg = housexml.compute_loyers()
            print(f"Loyers min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housexml.compute_surfaces()
            print(f"Surfaces min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housexml.compute_loyers_per_m2()
            print(f"Loyers par m² min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
        else:
            _, _, avg = housexml.compute_loyers_per_m2()
            print(f"{avg:.2f}")
    elif args.type == "json":
        housejson = HouseJson(args.path)
        housejson.load()
        if args.verbose:
            min, max, avg = housejson.compute_loyers()
            print(f"Loyers min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housejson.compute_surfaces()
            print(f"Surfaces min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housejson.compute_loyers_per_m2()
            print(f"Loyers par m² min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
        else:
            _, _, avg = housejson.compute_loyers_per_m2()
            print(f"{avg:.2f}")
    elif args.type == "xlsx":
        housexl = HouseExcel(args.path)
        housexl.load()
        if args.verbose:
            min, max, avg = housexl.compute_loyers()
            print(f"Loyers min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housexl.compute_surfaces()
            print(f"Surfaces min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
            min, max, avg = housexl.compute_loyers_per_m2()
            print(f"Loyers par m² min: {min:.0f}, max: {max:.0f}, avg: {avg:.1f}")
        else:
            _, _, avg = housexl.compute_loyers_per_m2()
            print(f"{avg:.2f}")
    elif args.type == "csv-pickle":
        housecsv = HouseCsv(args.path)
        housecsv.load()
        housepickle = HousePickle(args.path.replace(".csv", ".pickle"))
        housepickle.loyers = housecsv.loyers
        housepickle.surfaces = housecsv.surfaces
        housepickle.loyers_per_m2 = housecsv.loyers_per_m2
        housepickle.save()
    elif args.type == "csv-xml":
        housecsv = HouseCsv(args.path)
        housecsv.load()
        housexml = HouseXml(args.path.replace(".csv", ".xml"))
        housexml.loyers = housecsv.loyers
        housexml.surfaces = housecsv.surfaces
        housexml.loyers_per_m2 = housecsv.loyers_per_m2
        housexml.save()
    elif args.type == "csv-json":
        housecsv = HouseCsv(args.path)
        housecsv.load()
        housejson = HouseJson(args.path.replace(".csv", ".json"))
        housejson.loyers = housecsv.loyers
        housejson.surfaces = housecsv.surfaces
        housejson.loyers_per_m2 = housecsv.loyers_per_m2
        housejson.save()

    else:
        print("ERROR bad type")
        sys.exit(1)

    # Charger le fichier path et créer les listes loyers et surfaces
    # Rendez le séparateur paramétrable parse(path, separator)
    # Uniquement pour --type csv
    # Si verbose afficher min, max, moyenne des surfaces, loyers, loyers par m²
    # Si non verbose moyenne des loyers par m²

    # --type pickle
    # Faite tourner demofile.py 1 fois
    # HousePickle
    # Tester
    # Bonus : Ajouter la méthode save qui sauvegarde les data dans house.pickle
    # Bonus : Lire les data depuis HouseCsv et sauvegarder les data depuis HousePickle
    # Bonus : Reflechir à comment mutualiser HouseCsv et HousePickle

    # Inspirez vous de HousePickle pour créer HouseXml
    # Créer le type xml
    # Bonus : Créer un héritage : Créer la classe HouseSuper qui contient tous le code en commun
    # Bonus : Gérer l'héritage dans les 3 classes enfants
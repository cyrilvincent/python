import openpyxl

wb = openpyxl.load_workbook("data/house/house.xlsx")
sheet = wb["Sheet1"]
for row in range(1,sheet.max_row):
    value = sheet.cell(row, 2).value
    print(value)
